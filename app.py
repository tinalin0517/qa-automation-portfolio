import os
import io
import json
from flask import Flask, request, jsonify, render_template, send_file
from dotenv import load_dotenv
from opensearch_client import search_logs
from analyzer import (
    analyze_log, analyze_spec,
    deduplicate_cases, optimize_cases,
    build_code_index, extract_image_text
)

load_dotenv()
app = Flask(__name__)

# ══════════════════════════════════════════════════════════════
#  啟動時建立程式碼索引
# ══════════════════════════════════════════════════════════════

ZIP_PATH         = os.environ.get("PAYMENT_CODE_ZIP",  "sample-payment-service.zip")
PRICING_ZIP_PATH = os.environ.get("PRICING_CODE_ZIP", "sample-pricing-service.zip")
build_code_index(ZIP_PATH, PRICING_ZIP_PATH)


# ══════════════════════════════════════════════════════════════
#  基本路由
# ══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")


# ══════════════════════════════════════════════════════════════
#  Bug Analyzer
# ══════════════════════════════════════════════════════════════

@app.route("/analyze", methods=["POST"])
def analyze():
    data       = request.json
    keyword    = data.get("keyword", "")
    manual_log = data.get("errorLog", "")
    context    = data.get("context", "")
    time_range = data.get("timeRange", "")

    if not keyword and not manual_log:
        return jsonify({"error": "請提供關鍵字或 error log"}), 400

    log_content = manual_log
    raw_logs    = ""
    if keyword:
        try:
            raw_logs = search_logs(keyword, time_range=time_range)
            if raw_logs:
                if manual_log:
                    log_content = (
                        f"【使用者提供的 Error Log】\n{manual_log}\n\n"
                        f"【OpenSearch 查詢結果】\n{raw_logs}"
                    )
                else:
                    log_content = raw_logs
            elif not manual_log:
                return jsonify({"error": f"查無關鍵字：{keyword}"}), 404
        except Exception as e:
            return jsonify({"error": f"OpenSearch 查詢失敗：{str(e)}"}), 500

    try:
        # ★ 傳入 keyword 讓索引選對業務領域的 Controller
        result = analyze_log(log_content, context, keyword=keyword)
        result["rawLogs"] = raw_logs
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  Spec-to-Test
# ══════════════════════════════════════════════════════════════

@app.route("/spec-analyze", methods=["POST"])
def spec_analyze():
    spec_text = request.form.get("specText", "").strip()
    types_raw = request.form.get("types", "[]")
    bdd       = request.form.get("bdd",      "true")  == "true"
    priority  = request.form.get("priority", "true")  == "true"
    test_data = request.form.get("testData", "false") == "true"

    try:
        types = json.loads(types_raw)
    except Exception:
        types = []

    # ★ 多檔上傳：讀取所有 'files' key，同時相容舊版單檔 'file' key
    uploaded_files = request.files.getlist("files")
    single = request.files.get("file")
    if single and single.filename:
        uploaded_files = [single] + uploaded_files

    file_texts = []
    for uploaded in uploaded_files:
        if not uploaded or not uploaded.filename:
            continue
        ext = uploaded.filename.rsplit(".", 1)[-1].lower()
        try:
            if ext in ("txt", "md", "yaml", "yml"):
                text = uploaded.read().decode("utf-8", errors="ignore")
            elif ext == "docx":
                import docx
                doc  = docx.Document(io.BytesIO(uploaded.read()))
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            elif ext == "pdf":
                from pypdf import PdfReader
                reader = PdfReader(io.BytesIO(uploaded.read()))
                text   = "\n".join(page.extract_text() or "" for page in reader.pages)
            elif ext in ("png", "jpg", "jpeg", "webp"):
                media_type = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"
                text = extract_image_text(uploaded.read(), media_type)
            else:
                continue
            if text.strip():
                file_texts.append(f"【檔案：{uploaded.filename}】\n{text.strip()}")
        except Exception as e:
            return jsonify({"error": f"檔案 {uploaded.filename} 解析失敗：{str(e)}"}), 400

    file_text = "\n\n---\n\n".join(file_texts)
    combined = "\n\n".join(filter(None, [file_text, spec_text]))
    if not combined:
        return jsonify({"error": "請上傳規格文件或貼上規格內容"}), 400

    if not types:
        types = ["Happy Path", "Edge Cases", "Error Handling"]

    try:
        result = analyze_spec(combined, types, bdd=bdd, priority=priority, test_data=test_data)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  Test Case Deduplicator  ★ 新增
# ══════════════════════════════════════════════════════════════

@app.route("/dedup-analyze", methods=["POST"])
def dedup_analyze():
    cases_text = request.form.get("casesText", "").strip()
    file_text  = ""

    uploaded = request.files.get("file")
    if uploaded and uploaded.filename:
        ext = uploaded.filename.rsplit(".", 1)[-1].lower()
        try:
            if ext in ("txt", "md", "csv"):
                file_text = uploaded.read().decode("utf-8", errors="ignore")
            elif ext in ("xlsx", "xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()))
                ws = wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = "  ".join(str(c) for c in row if c)
                    if row_text.strip():
                        rows.append(row_text)
                file_text = "\n".join(rows)
        except Exception as e:
            return jsonify({"error": f"檔案解析失敗：{str(e)}"}), 400

    combined = "\n\n".join(filter(None, [file_text, cases_text]))
    if not combined:
        return jsonify({"error": "請上傳測試案例檔案或貼上內容"}), 400

    try:
        result = deduplicate_cases(combined)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  Test Case Optimizer  ★ 新增
# ══════════════════════════════════════════════════════════════

@app.route("/optimize-analyze", methods=["POST"])
def optimize_analyze():
    cases_text = request.form.get("casesText", "").strip()
    file_text  = ""

    uploaded = request.files.get("file")
    if uploaded and uploaded.filename:
        ext = uploaded.filename.rsplit(".", 1)[-1].lower()
        try:
            if ext in ("txt", "md", "csv"):
                file_text = uploaded.read().decode("utf-8", errors="ignore")
            elif ext in ("xlsx", "xls"):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()))
                ws = wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = "  ".join(str(c) for c in row if c)
                    if row_text.strip():
                        rows.append(row_text)
                file_text = "\n".join(rows)
        except Exception as e:
            return jsonify({"error": f"檔案解析失敗：{str(e)}"}), 400

    combined = "\n\n".join(filter(None, [file_text, cases_text]))
    if not combined:
        return jsonify({"error": "請上傳測試案例檔案或貼上內容"}), 400

    try:
        result = optimize_cases(combined)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ══════════════════════════════════════════════════════════════
#  Spec Export Excel（原有邏輯不變）
# ══════════════════════════════════════════════════════════════

@app.route("/spec-export-excel", methods=["POST"])
def spec_export_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime

    data = request.json
    if not data or not data.get("groups"):
        return jsonify({"error": "無測試資料"}), 400

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(color, bold=False, size=10, name="Calibri", italic=False):
        return Font(color=color, bold=bold, size=size, name=name, italic=italic)

    def border(color="2C2C31"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def align(h="left", v="top", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    C = dict(
        title_bg="085041", title_fg="FFFFFF",
        sub_bg="0A4535",   sub_fg="4ECB9A",
        hdr_bg="0D6652",   hdr_fg="FFFFFF",
        sep="0E0E10",
        row_a="1C1C1F",    row_b="141416",
        text="EEEEF2",     dim="7A7A8A",
        border="2C2C31",
        hp_bg="0D3326",  hp_fg="4ECB9A",
        ec_bg="2A1E06",  ec_fg="E8B84B",
        eh_bg="2D1212",  eh_fg="F08080",
        sec_bg="0D1F38", sec_fg="6AAEE8",
        oth_bg="202024", oth_fg="A0A0B0",
        p1_bg="3D1510",  p1_fg="F08080",
        p2_bg="2A1E06",  p2_fg="E8B84B",
        p3_bg="0D1F38",  p3_fg="6AAEE8",
    )
    GROUP_COLORS = {
        "success":   ("hp_bg",  "hp_fg"),
        "warning":   ("ec_bg",  "ec_fg"),
        "danger":    ("eh_bg",  "eh_fg"),
        "info":      ("sec_bg", "sec_fg"),
        "secondary": ("oth_bg", "oth_fg"),
    }
    PRIO_COLORS = {
        "P1": ("p1_bg", "p1_fg"),
        "P2": ("p2_bg", "p2_fg"),
        "P3": ("p3_bg", "p3_fg"),
    }

    groups    = data.get("groups", [])
    all_cases = [c for g in groups for c in g.get("cases", [])]
    total     = len(all_cases)
    p1_count  = sum(1 for c in all_cases if c.get("priority") == "P1")
    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "測試場景"
    ws.sheet_view.showGridLines = False

    COLS   = ["A", "B", "C", "D", "E", "F"]
    WIDTHS = [14,   16,   34,   58,   22,   9]
    HDRS   = ["編號", "測試類型", "名稱", "驗證步驟（Given / When / Then）", "Tags", "優先級"]

    for col, w in zip(COLS, WIDTHS):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 36
    c = ws["A1"]
    c.value     = "🧪  AI Test Assistant　｜　測試場景報告"
    c.fill      = fill(C["title_bg"])
    c.font      = Font(color=C["title_fg"], bold=True, size=15, name="Calibri")
    c.alignment = align("center", "center")
    for col in COLS[1:]:
        ws[f"{col}1"].fill = fill(C["title_bg"])

    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 18
    c = ws["A2"]
    c.value = (f"產生時間：{now_str}"
               f"　　總場景數：{total}"
               f"　　P1 高優先：{p1_count}"
               f"　　涵蓋類型：{len(groups)} 類")
    c.fill      = fill(C["sub_bg"])
    c.font      = Font(color=C["sub_fg"], size=9, name="Calibri")
    c.alignment = align("center", "center")
    for col in COLS[1:]:
        ws[f"{col}2"].fill = fill(C["sub_bg"])

    ws.row_dimensions[3].height = 5
    for col in COLS:
        ws[f"{col}3"].fill = fill(C["sep"])

    ws.row_dimensions[4].height = 24
    for ci, (h, col) in enumerate(zip(HDRS, COLS)):
        c = ws[f"{col}4"]
        c.value     = h
        c.fill      = fill(C["hdr_bg"])
        c.font      = font(C["hdr_fg"], bold=True, size=10)
        c.alignment = align("center", "center")
        c.border    = border()

    ws.auto_filter.ref = "A4:F4"
    ws.freeze_panes    = "A5"

    row = 5
    for group in groups:
        btype          = group.get("badgeType", "secondary")
        bg_key, fg_key = GROUP_COLORS.get(btype, ("oth_bg", "oth_fg"))
        gbg, gfg       = C[bg_key], C[fg_key]
        label          = group.get("badgeLabel", group.get("name", ""))

        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:F{row}")
        c           = ws[f"A{row}"]
        c.value     = f"▌  {group.get('name', '')}　{label}　（{len(group.get('cases', []))} 個場景）"
        c.fill      = fill(gbg)
        c.font      = font(gfg, bold=True, size=10)
        c.alignment = align("left", "center")
        for col in COLS[1:]:
            ws[f"{col}{row}"].fill = fill(gbg)
        row += 1

        for i, case in enumerate(group.get("cases", [])):
            row_bg            = C["row_a"] if i % 2 == 0 else C["row_b"]
            ws.row_dimensions[row].height = 52
            prio              = case.get("priority", "P2")
            pbg_key, pfg_key  = PRIO_COLORS.get(prio, ("p2_bg", "p2_fg"))
            pbg, pfg          = C[pbg_key], C[pfg_key]
            tags_str          = "  ·  ".join(case.get("tags", []))

            cells = [
                ("A", case.get("id", ""),    row_bg, C["dim"],  False, 9,  "center", False, "Courier New"),
                ("B", group.get("name", ""), gbg,    gfg,       True,  9,  "center", False, "Calibri"),
                ("C", case.get("name", ""),  row_bg, C["text"], True,  10, "left",   True,  "Calibri"),
                ("D", case.get("exp", ""),   row_bg, C["dim"],  False, 9,  "left",   True,  "Calibri"),
                ("E", tags_str,              row_bg, C["dim"],  False, 9,  "left",   True,  "Calibri"),
                ("F", prio,                  pbg,    pfg,       True,  11, "center", False, "Calibri"),
            ]
            for col, val, bg, fg, bold, size, halign, wrap_text, fname in cells:
                c           = ws[f"{col}{row}"]
                c.value     = val
                c.fill      = fill(bg)
                c.font      = Font(color=fg, bold=bold, size=size, name=fname)
                c.alignment = Alignment(horizontal=halign, vertical="top", wrap_text=wrap_text)
                c.border    = border()
            row += 1

        ws.row_dimensions[row].height = 5
        for col in COLS:
            ws[f"{col}{row}"].fill = fill(C["sep"])
        row += 1

    ws2 = wb.create_sheet("摘要")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14

    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 30
    c = ws2["A1"]
    c.value     = "測試場景統計摘要"
    c.fill      = fill(C["title_bg"])
    c.font      = Font(color=C["title_fg"], bold=True, size=13, name="Calibri")
    c.alignment = align("center", "center")
    ws2["B1"].fill = fill(C["title_bg"])

    summary_rows = [
        ("產生時間",   now_str),
        ("總場景數",   total),
        ("P1 高優先",  p1_count),
        ("P2 中優先",  sum(1 for c in all_cases if c.get("priority") == "P2")),
        ("P3 低優先",  sum(1 for c in all_cases if c.get("priority") == "P3")),
        ("涵蓋類型數", len(groups)),
    ]
    for i, (label, value) in enumerate(summary_rows):
        r      = i + 2
        row_bg = C["row_a"] if i % 2 == 0 else C["row_b"]
        ws2.row_dimensions[r].height = 20

        ca = ws2[f"A{r}"]
        ca.value     = label
        ca.fill      = fill(row_bg)
        ca.font      = font(C["dim"], bold=True)
        ca.alignment = align("left", "center")
        ca.border    = border()

        cb = ws2[f"B{r}"]
        cb.value     = value
        cb.fill      = fill(row_bg)
        cb.font      = font(C["text"])
        cb.alignment = align("center", "center")
        cb.border    = border()

    ws2.row_dimensions[len(summary_rows) + 3].height = 8
    start = len(summary_rows) + 4
    for i, group in enumerate(groups):
        r              = start + i
        btype          = group.get("badgeType", "secondary")
        bg_key, fg_key = GROUP_COLORS.get(btype, ("oth_bg", "oth_fg"))
        gbg2, gfg2     = C[bg_key], C[fg_key]
        ws2.row_dimensions[r].height = 20

        ca = ws2[f"A{r}"]
        ca.value     = group.get("name", "")
        ca.fill      = fill(gbg2)
        ca.font      = font(gfg2, bold=True)
        ca.alignment = align("left", "center")
        ca.border    = border()

        cb = ws2[f"B{r}"]
        cb.value     = len(group.get("cases", []))
        cb.fill      = fill(gbg2)
        cb.font      = font(gfg2)
        cb.alignment = align("center", "center")
        cb.border    = border()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="test-scenarios.xlsx"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(debug=debug, port=port)
